#!/usr/bin/env python3
"""
Excel Test File Generator for SinkerProject26
Generates test Excel files for Sales Forecast and Weekly Schedule uploads.
"""

import os
import sys
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is not installed. Please run: pip install openpyxl")
    sys.exit(1)


# Output directory
OUTPUT_DIR = Path("/output")

# Sample product data
PRODUCT_NAMES = [
    "可口可樂", "雪碧", "七喜", "芬達", "美粒果",
    "舒跑", "寶礦力", "黑松沙士", "茶裏王", "純喫茶",
    "統一麥香", "愛之味", "維大力", "蘋果西打", "金車麥根沙士",
    "冬瓜茶", "青草茶", "烏龍茶", "綠茶", "紅茶"
]

CATEGORIES = ["飲料類", "零食類", "日用品", "食品類", "調味品"]
SPECS = ["330ml罐裝", "500ml寶特瓶", "1500ml大瓶", "150g袋裝", "250g盒裝"]
WAREHOUSE_LOCATIONS = ["A區", "B區", "C區", "D區", "E區", "F區"]


def apply_excel_styling(ws):
    """Apply common styling to Excel worksheet."""
    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Freeze first row
    ws.freeze_panes = "A2"

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_product_code(index):
    """Generate product code in format P001, P002, ..., P100."""
    return f"P{index:03d}"


# =============================================================================
# Sales Forecast Excel Generators
# =============================================================================

def create_forecast_ok():
    """Generate valid sales forecast Excel with 10 rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Forecast"

    # Header (fixed order)
    headers = ["中類名稱", "貨品規格", "品號", "品名", "庫位", "箱數小計"]
    ws.append(headers)

    # Data rows (10 normal rows)
    for i in range(1, 11):
        row = [
            CATEGORIES[i % len(CATEGORIES)],
            SPECS[i % len(SPECS)],
            generate_product_code(i),
            PRODUCT_NAMES[i % len(PRODUCT_NAMES)],
            WAREHOUSE_LOCATIONS[i % len(WAREHOUSE_LOCATIONS)],
            float(Decimal(str(100 + i * 25)) + Decimal("0.50"))
        ]
        ws.append(row)

    # Format quantity column
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 6).number_format = '0.00'

    apply_excel_styling(ws)

    filepath = OUTPUT_DIR / "upload_forecast_ok.xlsx"
    wb.save(filepath)
    print(f"✅ Created: {filepath}")


def create_forecast_missing_required():
    """Generate Excel with missing required fields (product_code, quantity)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Forecast"

    headers = ["中類名稱", "貨品規格", "品號", "品名", "庫位", "箱數小計"]
    ws.append(headers)

    # Row 1: Missing product_code
    ws.append([CATEGORIES[0], SPECS[0], "", PRODUCT_NAMES[0], WAREHOUSE_LOCATIONS[0], 100.50])

    # Row 2: Missing quantity
    ws.append([CATEGORIES[1], SPECS[1], generate_product_code(2), PRODUCT_NAMES[1], WAREHOUSE_LOCATIONS[1], None])

    # Row 3: Both missing
    ws.append([CATEGORIES[2], SPECS[2], "", PRODUCT_NAMES[2], WAREHOUSE_LOCATIONS[2], None])

    # Row 4: Product code is blank (spaces only)
    ws.append([CATEGORIES[3], SPECS[3], "   ", PRODUCT_NAMES[3], WAREHOUSE_LOCATIONS[3], 50.00])

    apply_excel_styling(ws)

    filepath = OUTPUT_DIR / "upload_forecast_missing_required.xlsx"
    wb.save(filepath)
    print(f"✅ Created: {filepath}")


def create_forecast_bad_types():
    """Generate Excel with bad data types (negative, zero, text)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Forecast"

    headers = ["中類名稱", "貨品規格", "品號", "品名", "庫位", "箱數小計"]
    ws.append(headers)

    # Row 1: Negative quantity
    ws.append([CATEGORIES[0], SPECS[0], generate_product_code(1), PRODUCT_NAMES[0], WAREHOUSE_LOCATIONS[0], -50.00])

    # Row 2: Zero quantity (not allowed for forecast)
    ws.append([CATEGORIES[1], SPECS[1], generate_product_code(2), PRODUCT_NAMES[1], WAREHOUSE_LOCATIONS[1], 0.00])

    # Row 3: Text quantity
    ws.append([CATEGORIES[2], SPECS[2], generate_product_code(3), PRODUCT_NAMES[2], WAREHOUSE_LOCATIONS[2], "abc"])

    # Row 4: Very large negative
    ws.append([CATEGORIES[3], SPECS[3], generate_product_code(4), PRODUCT_NAMES[3], WAREHOUSE_LOCATIONS[3], -9999.99])

    apply_excel_styling(ws)

    filepath = OUTPUT_DIR / "upload_forecast_bad_types.xlsx"
    wb.save(filepath)
    print(f"✅ Created: {filepath}")


def create_forecast_duplicates():
    """Generate Excel with duplicate product codes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Forecast"

    headers = ["中類名稱", "貨品規格", "品號", "品名", "庫位", "箱數小計"]
    ws.append(headers)

    # Same product code repeated 5 times
    for i in range(5):
        ws.append([
            CATEGORIES[i % len(CATEGORIES)],
            SPECS[i % len(SPECS)],
            generate_product_code(1),  # Always P001
            PRODUCT_NAMES[i % len(PRODUCT_NAMES)],
            WAREHOUSE_LOCATIONS[i % len(WAREHOUSE_LOCATIONS)],
            float(100 + i * 10)
        ])

    # Different product codes
    for i in range(5, 10):
        ws.append([
            CATEGORIES[i % len(CATEGORIES)],
            SPECS[i % len(SPECS)],
            generate_product_code(i),
            PRODUCT_NAMES[i % len(PRODUCT_NAMES)],
            WAREHOUSE_LOCATIONS[i % len(WAREHOUSE_LOCATIONS)],
            float(100 + i * 10)
        ])

    apply_excel_styling(ws)

    filepath = OUTPUT_DIR / "upload_forecast_duplicates.xlsx"
    wb.save(filepath)
    print(f"✅ Created: {filepath}")


def create_forecast_large_5k():
    """Generate large Excel with 5000 rows for performance testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Forecast"

    headers = ["中類名稱", "貨品規格", "品號", "品名", "庫位", "箱數小計"]
    ws.append(headers)

    # Generate 5000 rows
    for i in range(1, 5001):
        row = [
            CATEGORIES[i % len(CATEGORIES)],
            SPECS[i % len(SPECS)],
            generate_product_code((i % 100) + 1),  # Cycle through P001-P100
            PRODUCT_NAMES[i % len(PRODUCT_NAMES)],
            WAREHOUSE_LOCATIONS[i % len(WAREHOUSE_LOCATIONS)],
            float(Decimal(str(50 + (i % 500))) + Decimal("0.25"))
        ]
        ws.append(row)

    apply_excel_styling(ws)

    filepath = OUTPUT_DIR / "upload_forecast_large_5k.xlsx"
    wb.save(filepath)
    print(f"✅ Created: {filepath} (5000 rows)")


def generate_forecast_excels():
    """Generate all sales forecast test Excel files."""
    print("\n📊 Generating Sales Forecast Excel files...")
    create_forecast_ok()
    create_forecast_missing_required()
    create_forecast_bad_types()
    create_forecast_duplicates()
    create_forecast_large_5k()


# =============================================================================
# Weekly Schedule Excel Generators
# =============================================================================

def create_schedule_ok():
    """Generate valid weekly schedule Excel with 20 rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Schedule"

    # Header (dynamic order - Chinese names for detection)
    headers = ["需求日期", "品號", "品名", "庫位", "箱數小計"]
    ws.append(headers)

    # Start from Monday 2026-03-02
    start_date = date(2026, 3, 2)  # Monday

    # Generate 20 rows (covering 1 week, multiple products per day)
    for i in range(20):
        day_offset = i % 7
        demand_date = start_date + timedelta(days=day_offset)

        row = [
            demand_date,  # Excel will format as date
            generate_product_code((i % 100) + 1),
            PRODUCT_NAMES[i % len(PRODUCT_NAMES)],
            WAREHOUSE_LOCATIONS[i % len(WAREHOUSE_LOCATIONS)],
            float(Decimal(str(80 + i * 15)) + Decimal("0.75"))
        ]
        ws.append(row)

    # Format date column
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).number_format = 'YYYY-MM-DD'

    # Format quantity column
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 5).number_format = '0.00'

    apply_excel_styling(ws)

    filepath = OUTPUT_DIR / "upload_schedule_ok.xlsx"
    wb.save(filepath)
    print(f"✅ Created: {filepath}")


def create_schedule_missing_required():
    """Generate Excel with missing required fields."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Schedule"

    headers = ["需求日期", "品號", "品名", "庫位", "箱數小計"]
    ws.append(headers)

    base_date = date(2026, 3, 2)

    # Row 1: Missing demand_date
    ws.append([None, generate_product_code(1), PRODUCT_NAMES[0], WAREHOUSE_LOCATIONS[0], 100.50])

    # Row 2: Missing product_code
    ws.append([base_date, "", PRODUCT_NAMES[1], WAREHOUSE_LOCATIONS[1], 100.50])

    # Row 3: Missing product_name
    ws.append([base_date + timedelta(days=1), generate_product_code(3), "", WAREHOUSE_LOCATIONS[2], 100.50])

    # Row 4: Missing warehouse_location
    ws.append([base_date + timedelta(days=2), generate_product_code(4), PRODUCT_NAMES[3], "", 100.50])

    # Row 5: Missing quantity
    ws.append([base_date + timedelta(days=3), generate_product_code(5), PRODUCT_NAMES[4], WAREHOUSE_LOCATIONS[4], None])

    apply_excel_styling(ws)

    filepath = OUTPUT_DIR / "upload_schedule_missing_required.xlsx"
    wb.save(filepath)
    print(f"✅ Created: {filepath}")


def create_schedule_bad_types():
    """Generate Excel with bad data types (invalid dates, negative quantity)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Schedule"

    headers = ["需求日期", "品號", "品名", "庫位", "箱數小計"]
    ws.append(headers)

    base_date = date(2026, 3, 2)

    # Row 1: Invalid date (text)
    ws.append(["abc", generate_product_code(1), PRODUCT_NAMES[0], WAREHOUSE_LOCATIONS[0], 100.50])

    # Row 2: Negative quantity (not allowed)
    ws.append([base_date, generate_product_code(2), PRODUCT_NAMES[1], WAREHOUSE_LOCATIONS[1], -50.00])

    # Row 3: Text quantity
    ws.append([base_date + timedelta(days=1), generate_product_code(3), PRODUCT_NAMES[2], WAREHOUSE_LOCATIONS[2], "xyz"])

    # Row 4: Invalid date format
    ws.append(["2026/13/45", generate_product_code(4), PRODUCT_NAMES[3], WAREHOUSE_LOCATIONS[3], 100.00])

    apply_excel_styling(ws)

    filepath = OUTPUT_DIR / "upload_schedule_bad_types.xlsx"
    wb.save(filepath)
    print(f"✅ Created: {filepath}")


def create_schedule_duplicates():
    """Generate Excel with duplicate product+date combinations."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Schedule"

    headers = ["需求日期", "品號", "品名", "庫位", "箱數小計"]
    ws.append(headers)

    base_date = date(2026, 3, 2)

    # Same product code + date repeated 5 times
    for i in range(5):
        ws.append([
            base_date,
            generate_product_code(1),  # Always P001
            PRODUCT_NAMES[0],
            WAREHOUSE_LOCATIONS[i % len(WAREHOUSE_LOCATIONS)],
            float(100 + i * 10)
        ])

    # Different combinations
    for i in range(5, 10):
        ws.append([
            base_date + timedelta(days=i % 7),
            generate_product_code(i),
            PRODUCT_NAMES[i % len(PRODUCT_NAMES)],
            WAREHOUSE_LOCATIONS[i % len(WAREHOUSE_LOCATIONS)],
            float(100 + i * 10)
        ])

    apply_excel_styling(ws)

    filepath = OUTPUT_DIR / "upload_schedule_duplicates.xlsx"
    wb.save(filepath)
    print(f"✅ Created: {filepath}")


def create_schedule_large_5k():
    """Generate large Excel with 5000 rows for performance testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Schedule"

    headers = ["需求日期", "品號", "品名", "庫位", "箱數小計"]
    ws.append(headers)

    base_date = date(2026, 3, 2)

    # Generate 5000 rows
    for i in range(5000):
        demand_date = base_date + timedelta(days=i % 30)  # Spread over 30 days

        row = [
            demand_date,
            generate_product_code((i % 100) + 1),
            PRODUCT_NAMES[i % len(PRODUCT_NAMES)],
            WAREHOUSE_LOCATIONS[i % len(WAREHOUSE_LOCATIONS)],
            float(Decimal(str(50 + (i % 500))) + Decimal("0.25"))
        ]
        ws.append(row)

    apply_excel_styling(ws)

    filepath = OUTPUT_DIR / "upload_schedule_large_5k.xlsx"
    wb.save(filepath)
    print(f"✅ Created: {filepath} (5000 rows)")


def generate_schedule_excels():
    """Generate all weekly schedule test Excel files."""
    print("\n📅 Generating Weekly Schedule Excel files...")
    create_schedule_ok()
    create_schedule_missing_required()
    create_schedule_bad_types()
    create_schedule_duplicates()
    create_schedule_large_5k()


# =============================================================================
# Main
# =============================================================================

def main():
    """Main entry point."""
    print("=" * 60)
    print("Excel Test File Generator for SinkerProject26")
    print("=" * 60)

    # Create output directory if it doesn't exist
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Generate all test files
    generate_forecast_excels()
    generate_schedule_excels()

    print("\n" + "=" * 60)
    print("✅ All test Excel files generated successfully!")
    print(f"📁 Output directory: {OUTPUT_DIR}")
    print("=" * 60)

    # List generated files
    files = sorted(OUTPUT_DIR.glob("*.xlsx"))
    print(f"\n📋 Generated {len(files)} files:")
    for f in files:
        size_kb = f.stat().st_size / 1024
        print(f"  - {f.name} ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
